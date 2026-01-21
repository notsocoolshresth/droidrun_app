"""
Job Tracker Module
Manages job applications in Excel format for tracking and status updates.
"""
import os
from datetime import datetime
from typing import Dict, List, Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class JobTracker:
    """Tracks job applications in an Excel spreadsheet."""
    
    def __init__(self, excel_path: str):
        """
        Initialize the job tracker.
        
        Args:
            excel_path: Path to the Excel file for tracking
        """
        self.excel_path = excel_path
        self.workbook = None
        self.worksheet = None
        self._initialize_workbook()
    
    def _initialize_workbook(self):
        """Initialize or load the Excel workbook."""
        if os.path.exists(self.excel_path):
            try:
                self.workbook = openpyxl.load_workbook(self.excel_path)
                self.worksheet = self.workbook.active
                print(f"✅ Loaded existing job tracker: {self.excel_path}")
            except Exception as e:
                print(f"⚠️ Error loading workbook, creating new one: {e}")
                self._create_new_workbook()
        else:
            self._create_new_workbook()
    
    def _create_new_workbook(self):
        """Create a new workbook with headers."""
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)
        
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Job Applications"
        
        # Define headers
        headers = [
            "Application ID",
            "Date Applied",
            "Platform",
            "Company",
            "Job Title",
            "Location",
            "Job URL",
            "Status",
            "Experience Required",
            "Job Type",
            "Salary Range",
            "Key Skills",
            "Application Method",
            "Last Updated",
            "Notes"
        ]
        
        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        column_widths = [15, 20, 12, 25, 30, 20, 50, 15, 18, 12, 15, 30, 18, 20, 40]
        for col, width in enumerate(column_widths, start=1):
            self.worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        self.workbook.save(self.excel_path)
        print(f"✅ Created new job tracker: {self.excel_path}")
    
    def add_job_application(self, job_data: Dict[str, str]) -> bool:
        """
        Add a new job application to the tracker.
        
        Args:
            job_data: Dictionary containing job information
            
        Returns:
            True if successful, False otherwise
        """
        try:
            next_row = self.worksheet.max_row + 1
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Generate application ID
            app_id = f"{job_data.get('platform', 'UNK')[:3].upper()}-{timestamp.replace(' ', 'T').replace(':', '').replace('-', '')}"
            
            # Prepare row data
            row_data = [
                app_id,
                timestamp,
                job_data.get('platform', 'Unknown'),
                job_data.get('company', 'Unknown'),
                job_data.get('job_title', 'Unknown'),
                job_data.get('location', 'Unknown'),
                job_data.get('job_url', ''),
                job_data.get('status', 'Applied'),
                job_data.get('experience', '0-1 years'),
                job_data.get('job_type', 'Internship'),
                job_data.get('salary', 'Not specified'),
                job_data.get('skills', ''),
                job_data.get('application_method', 'Automated'),
                timestamp,
                job_data.get('notes', '')
            ]
            
            # Write row
            for col, value in enumerate(row_data, start=1):
                self.worksheet.cell(row=next_row, column=col, value=value)
            
            # Color-code status
            status_cell = self.worksheet.cell(row=next_row, column=8)
            status = job_data.get('status', 'Applied').lower()
            if status == 'applied':
                status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            elif status in ['rejected', 'closed']:
                status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            elif status in ['interview', 'shortlisted']:
                status_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            
            self.workbook.save(self.excel_path)
            print(f"✅ Added job application: {job_data.get('job_title')} at {job_data.get('company')}")
            return True
            
        except Exception as e:
            print(f"❌ Error adding job application: {e}")
            return False
    
    def update_job_status(self, app_id: str, new_status: str, notes: str = ""):
        """
        Update the status of an existing job application.
        
        Args:
            app_id: Application ID to update
            new_status: New status value
            notes: Additional notes
        """
        try:
            for row in range(2, self.worksheet.max_row + 1):
                if self.worksheet.cell(row=row, column=1).value == app_id:
                    self.worksheet.cell(row=row, column=8, value=new_status)
                    self.worksheet.cell(row=row, column=14, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    if notes:
                        existing_notes = self.worksheet.cell(row=row, column=15).value or ""
                        updated_notes = f"{existing_notes}\n{datetime.now().strftime('%Y-%m-%d')}: {notes}".strip()
                        self.worksheet.cell(row=row, column=15, value=updated_notes)
                    
                    self.workbook.save(self.excel_path)
                    print(f"✅ Updated status for {app_id}: {new_status}")
                    return True
            
            print(f"⚠️ Application ID {app_id} not found")
            return False
            
        except Exception as e:
            print(f"❌ Error updating job status: {e}")
            return False
    
    def check_already_applied(self, company: str, job_title: str, platform: str) -> bool:
        """
        Check if already applied to a specific job.
        
        Args:
            company: Company name
            job_title: Job title
            platform: Platform name
            
        Returns:
            True if already applied, False otherwise
        """
        try:
            for row in range(2, self.worksheet.max_row + 1):
                existing_company = self.worksheet.cell(row=row, column=4).value or ""
                existing_title = self.worksheet.cell(row=row, column=5).value or ""
                existing_platform = self.worksheet.cell(row=row, column=3).value or ""
                
                if (existing_company.lower() == company.lower() and 
                    existing_title.lower() == job_title.lower() and
                    existing_platform.lower() == platform.lower()):
                    return True
            
            return False
            
        except Exception as e:
            print(f"❌ Error checking application history: {e}")
            return False
    
    def get_application_stats(self) -> Dict[str, int]:
        """
        Get statistics about job applications.
        
        Returns:
            Dictionary with application statistics
        """
        stats = {
            'total': 0,
            'applied': 0,
            'interview': 0,
            'rejected': 0,
            'pending': 0
        }
        
        try:
            for row in range(2, self.worksheet.max_row + 1):
                stats['total'] += 1
                status = (self.worksheet.cell(row=row, column=8).value or "").lower()
                
                if 'applied' in status:
                    stats['applied'] += 1
                elif 'interview' in status or 'shortlist' in status:
                    stats['interview'] += 1
                elif 'reject' in status:
                    stats['rejected'] += 1
                else:
                    stats['pending'] += 1
            
            return stats
            
        except Exception as e:
            print(f"❌ Error getting stats: {e}")
            return stats
    
    def close(self):
        """Close and save the workbook."""
        try:
            if self.workbook:
                self.workbook.save(self.excel_path)
                self.workbook.close()
                print(f"✅ Saved and closed job tracker")
        except Exception as e:
            print(f"❌ Error closing workbook: {e}")
